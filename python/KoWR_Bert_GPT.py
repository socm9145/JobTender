# krWordRank base code

from krwordrank.word import KRWordRank
from keybert import KeyBERT
from konlpy.tag import Okt

import openai, time
from openpyxl import load_workbook
from openpyxl import Workbook

okt = Okt()
TC = 6 # 인재상 문장을 반복하는 횟수. 적은 문장에서 키워드를 찾기 위해 반복함.

# 메인, 서브 키워드에서 잘못 나온 결과를 찾아 직접 삽입. 해당 리스트 내부에 있는 단어가 나올 경우 메인, 서브 키워드에 포함시키기 않는다.
wrong_word = ["통해", "자신", "각자", "자기", "대안", "가지",
              "서로", "조성", "급변", "대한", "제시", "인재",
              "전체", "기존", "자발", "마인드", "위해", "여기",
              "다합", "안주", "통한", "감히", "중시", "요약",
              "두려움", "주의", "우선", "사람"]
# -*- coding: utf-8 -*-

main_all = set()
sub_all = set()

# KoWr 모델
def KoWordRankModule(snts):
    # 시간 측정 시작
    start = time.time()
    result = {"time": 0, "keyword":{}} # 모델을 돌리는데 걸린 시간과 발생한 키워드를 삽입

    texts = []

    # 전처리
    for _ in range(TC):
        for snt in snts:
            # 불러온 인재상 문장에서 개행과 따옴표를 제거
            info = '\'' + snt.replace("\n", "").replace("\'", "\\'") + '\','

            if len(info) != 0:
                texts.append(info)

    # 함수 적용
    wordrank_extractor = KRWordRank(
        min_count=5,  # 단어의 최소 출현 빈도수 (default: 5)
        max_length=10,  # 단어의 최대 길이 (default: 10)
        verbose=True
    )

    # 키워드 추출
    keywords, rank, graph = wordrank_extractor.extract(texts, num_keywords=30)
    vmax = 0

    # 결과 출력. word에 단어, r에 해당 단어의 가중치를 출력
    for word, r in sorted(keywords.items(), key=lambda x: x[1], reverse=True)[:30]:
        # koWR 모델의 결과에서 명사 추출
        name = okt.nouns(str(word).replace("\"", "").replace("\'", ""))
        
        # 추출된 명사가 있을 경우
        if name:
            val = round(r, 5)
            vmax = max(val, vmax) # 명사 가중치의 최대값 저장
            result["keyword"][name[0]] = val # 명사 저장
    
    # 모든 값을 명사 가중치의 최대값으로 나누어 0 ~ 1 사이로 변환
    for k in result["keyword"].keys():
        result["keyword"][k] = round(val / vmax, 5)
    
    # 시간 측정 종료
    end = time.time()
    
    # 시간 측정치 저장
    result["time"] = round(end - start, 5)

    return result

# -*- coding: utf-8 -*-
def KeyBert(texts):
    # 시간 측정 시작
    start = time.time()
    result = {"time": 0, "keyword": {}}
    doc = "" # 인재상 문장을 모은 하나의 문장
    
    # 인재상 문장을 모은 하나의 문장 생성
    for _ in range(TC):
        for text in texts:
            doc += text
    
    # 모델 볼러오기
    kw_model = KeyBERT()
    
    # 모델 튜닝 및 결과 반환.
    # (중요한 문장, 가중치) 형태로 반환
    keywords_mmr = kw_model.extract_keywords(doc, keyphrase_ngram_range=(2, 4), use_mmr=True, top_n=20, diversity=0.3)
    res_dict = {}

    for text in keywords_mmr:
        for word in okt.nouns(text[0]): # 중요한 문장을 명사로 쪼개는 작업
            if len(word) > 1: # 명사의 길이가 1보다 클 경우 작동 (길이가 1인 경우 의미 없는 단어가 많이 나오기 때문에 설명)
                if word in res_dict: # 같은 명사가 저장되는 경우 가중치가 큰 값으로 변경
                    res_dict[word] = max(text[1], res_dict[word])

                else:
                    res_dict[word] = text[1]

    # 단어 가중치 중 최대값 추출
    vmax = 0
    for key, value in res_dict.items():
        vmax = max(value, vmax)

    # 단어 가중치들을 최대값으로 나누어 가중치를 0 ~ 1로 조정
    for key, value in res_dict.items():
        res_dict[key] = round(value / vmax, 5)

    # 가중치 순서로 정렬
    d2 = sorted(res_dict.items(), key=lambda x: x[1], reverse=True)

    # 가중치 순서로 정렬
    for key in d2:result["keyword"][key[0]] = str(round(key[1], 5))

    # 시간 측정 종료
    end = time.time()
    result["time"] = round(end - start, 5)
    
    # 결과 반환
    return result

# 명사 추출 모델
def SenNouns(texts):
    # 시간 측정 시작
    start = time.time()
    result = {"time": 0, "keyword": {}}

    data = {} # (명사, 빈도수)를 담은 딕셔너리 변수

    # 인재상 문장에서 명사를 분해한 뒤 삽입
    for text in texts:
        nouns = okt.nouns(text)

        for noun in nouns:
            if noun in data:
                data[noun] += 1

            else:
                data[noun] = 1

    # 시간 측정 종료
    end = time.time()
    result["time"] = round(end - start, 5)

    # 명사 빈도수에 따라 정렬
    data_list = []

    for key, value in data.items():
        data_list.append([key, value])

    data_list.sort(key=lambda x: x[1], reverse=True)

    for dl in data_list:
        result["keyword"][dl[0]] = dl[1]

    # 빈도 최대값 변수
    vmax = 0

    # 빈도 최대값 추출
    for key, value in result["keyword"].items():
        vmax = max(value, vmax)

    # 모든 데이터를 최대값으로 나누어 0~1 가중치로 변경
    for key, value in result["keyword"].items():
        result["keyword"][key] = round(value / vmax, 5)

    return result

def GPT(texts):
    # 시간 측정 시작
    start = time.time()
    result = {"time": 0, "keyword": {}}

    # set api key
    OPENAI_API_KEY = "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"
    openai.api_key = OPENAI_API_KEY
    flag = True

    # 쿼리 정의하기
    while(flag):
        flag = False
        queries = ""

        # 정의한 쿼리
        for text in texts:
            queries += "\"{}\"를 한단어로 요약해줘. ".format(text)

        # 모델 - GPT 3.5 Turbo 선택
        model = "gpt-3.5-turbo"

        # 메시지 설정하기
        messages = [
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": queries}
        ]

        # ChatGPT API 호출하기
        response = openai.ChatCompletion.create(
            model=model,
            messages=messages
        )
        answer = response['choices'][0]['message']['content']

        # 키워드에서 명사 추출
        keywords = list(set(okt.nouns(answer)))

        # 추출한 명사의 크기가 1보다 작은 경우 GPT 재가동
        for keyword in keywords:
            if len(keyword) <= 1:
                flag = True
                break

        # 아무것도 추출되지 않은 경우 GPT 재가동
        if len(keywords) == 0:
            flag = True

        # 재가동 로직
        if flag:
            print("sleep...")
            time.sleep(1)

    # 추출된 명사에 가중치 1을 할당
    for keyword in keywords:
        if keyword not in wrong_word:
            result["keyword"][keyword] = 1

    # 최댓값 추출
    '''
    vmax = 0

    for key, value in result["keyword"].items():
        vmax = max(value, vmax)

    # 최댓값으로 나누기
    for key, value in result["keyword"].items():
        result["keyword"][key] = round(value / vmax, 5)
    '''

    # 시간 측정 종료
    end = time.time()
    result["time"] = round(end - start, 5)

    return result

def CUSTOM_MODEL_01(KoWR_result, Bert_result, SenNouns_result, wrong_word, V):
    result = {"time":{}, "keyword":{}}
    start = time.time()

    # Bert Base
    for key, value in Bert_result["keyword"].items():
        if key in list(KoWR_result["keyword"].keys()):
            result["keyword"][key] = float(value)

    # Add KoWR
    for key, value in KoWR_result["keyword"].items():
        if key in result["keyword"]:
            result["keyword"][key] += (value * V)

    # Delete frequent noun
    '''
    noun = 0

    for key, value in SenNouns_result["keyword"].items():
        if value == 1:
            noun = key
            break

    if noun != 0:
        for key, value in result["keyword"].items():
            if key == noun:
                del result["keyword"][key]
                break
    '''
    # desc order
    D1 = sorted(result["keyword"].items(), key=lambda x: x[1], reverse=True)
    result["keyword"] = {}

    for d1 in D1:
        # delete wrong word
        if d1[0] not in wrong_word:
            result["keyword"][d1[0]] = d1[1]

    end = time.time()
    result["time"] = round(end - start, 5)

    return result

def WriteExcel(path, input_name, output_name):
    input_path = "{}/{}".format(path, input_name)
    output_path = "{}/{}".format(path, output_name)

    workbook_input = load_workbook(filename=input_path)
    worksheet_input = workbook_input['Sheet1']

    workbook_output = Workbook()
    worksheet_output = workbook_output.create_sheet("Sheet1")
    worksheet_output = workbook_output.active

    worksheet_output['A1'] = "인덱스"
    worksheet_output['B1'] = "회사 이름"
    worksheet_output['C1'] = "메인 키워드"
    worksheet_output['D1'] = "서브 키워드"

    result = []
    output_idx = 2

    for i in range(1, 10000):
        code_val = worksheet_input["A{}".format(i)].value
        code_name = worksheet_input["B{}".format(i)].value
        code_info = worksheet_input["C{}".format(i)].value

        if code_val is not None and int(code_val) == 0:
            break

        if code_info is not None:
            if code_val is not None:
                result.append([int(code_val), code_name, [code_info.replace("\n", "")]])

            else:
                result[-1][2].append(code_info)

    for res in result:
        KoWR_result = KoWordRankModule(res[2])
        Bert_result = KeyBert(res[2])
        SenNouns_result = SenNouns(res[2])
        GPT_result = GPT(res[2])
        custom_model_01 = CUSTOM_MODEL_01(KoWR_result, Bert_result, SenNouns_result, wrong_word, 0.5)

        main_keys = []
        sub_keys = []

        print("Main keyword")
        for key, value in GPT_result["keyword"].items():
            main_keys.append(key)
            main_all.add(key)
        print(main_keys)

        print("Sub keyword")
        for key, value in custom_model_01["keyword"].items():
            sub_keys.append(key)
            sub_all.add(key)
        print(sub_keys)

        # write
        worksheet_output["A{}".format(output_idx)] = str(res[0])
        worksheet_output["B{}".format(output_idx)] = str(res[1])

        midx = output_idx
        for i in range(len(main_keys)):
            worksheet_output["C{}".format(midx)] = str(main_keys[i])
            midx += 1

        sidx = output_idx
        for i in range(len(sub_keys)):
            worksheet_output["D{}".format(sidx)] = str(sub_keys[i])
            sidx += 1

        output_idx += max(len(main_keys), len(sub_keys))

        print("END")

    workbook_output.save(output_path)
    print(len(main_all))
    print(main_all)

    print(len(sub_all))
    print(sub_all)

if __name__ == '__main__':
    path = "../data/info01.txt"
    beginTime = time.time()

    WriteExcel("C:/Users/SSAFY/Desktop/excel", "input11_20.xlsx", "output11_20.xlsx")

    endTime = time.time()
    print(round(endTime - beginTime))