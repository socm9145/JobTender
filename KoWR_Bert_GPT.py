# krWordRank base code

from krwordrank.word import KRWordRank
from keybert import KeyBERT
from konlpy.tag import Okt

import openai, time
from openpyxl import load_workbook
from openpyxl import Workbook

okt = Okt()
TC = 6
wrong_word = ["통해", "자신", "각자", "자기", "대안", "가지",
              "서로", "조성", "급변", "대한", "제시", "인재",
              "전체", "기존", "자발", "마인드", "위해", "여기",
              "다합", "안주", "통한", "감히", "중시", "요약",
              "두려움", "주의", "우선"]
# -*- coding: utf-8 -*-

main_all = set()
sub_all = set()

def KoWordRankModule(snts):
    start = time.time()
    result = {"time": 0, "keyword":{}}

    # 파일 불러오기
    texts = []

    # 전처리
    for _ in range(TC):
        for snt in snts:
            info = '\'' + snt.replace("\n", "").replace("\'", "\\'") + '\','
            if len(info) != 0:
                texts.append(info)

    # 함수 적용
    wordrank_extractor = KRWordRank(
        min_count=5,  # 단어의 최소 출현 빈도수 (default: 5)
        max_length=10,  # 단어의 최대 길이 (default: 10)
        verbose=True
    )

    # 키워드 추출
    keywords, rank, graph = wordrank_extractor.extract(texts, num_keywords=30)
    vmax = 0

    # 결과 출력
    for word, r in sorted(keywords.items(), key=lambda x: x[1], reverse=True)[:30]:
        name = okt.nouns(str(word).replace("\"", "").replace("\'", ""))
        if name:
            val = round(r, 5)
            vmax = max(val, vmax)
            result["keyword"][name[0]] = round(val / vmax, 5)

    end = time.time()
    result["time"] = round(end - start, 5)

    return result

# -*- coding: utf-8 -*-
def KeyBert(texts):
    start = time.time()
    result = {"time": 0, "keyword": {}}
    doc = ""

    for _ in range(TC):
        for text in texts:
            doc += text

    kw_model = KeyBERT()

    keywords_mmr = kw_model.extract_keywords(doc, keyphrase_ngram_range=(2, 4), use_mmr=True, top_n=20, diversity=0.3)
    res_dict = {}

    for text in keywords_mmr:
        for word in okt.nouns(text[0]):
            if len(word) > 1:
                if word in res_dict:
                    res_dict[word] = max(text[1], res_dict[word])

                else:
                    res_dict[word] = text[1]
    vmax = 0
    for key, value in res_dict.items():
        vmax = max(value, vmax)

    for key, value in res_dict.items():
        res_dict[key] = round(value / vmax, 5)

    d2 = sorted(res_dict.items(), key=lambda x: x[1], reverse=True)

    for key in d2:
        result["keyword"][key[0]] = str(round(key[1], 5))

    end = time.time()
    result["time"] = round(end - start, 5)

    return result

def SenNouns(texts):
    start = time.time()
    result = {"time": 0, "keyword": {}}

    data = {}

    for text in texts:
        nouns = okt.nouns(text)

        for noun in nouns:
            if noun in data:
                data[noun] += 1

            else:
                data[noun] = 1

    end = time.time()
    result["time"] = round(end - start, 5)

    data_list = []

    for key, value in data.items():
        data_list.append([key, value])

    data_list.sort(key=lambda x: x[1], reverse=True)

    for dl in data_list:
        result["keyword"][dl[0]] = dl[1]

    vmax = 0

    for key, value in result["keyword"].items():
        vmax = max(value, vmax)

    for key, value in result["keyword"].items():
        result["keyword"][key] = round(value / vmax, 5)

    return result

def GPT(texts):
    start = time.time()
    result = {"time": 0, "keyword": {}}

    # set api key
    OPENAI_API_KEY = "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"
    openai.api_key = OPENAI_API_KEY
    flag = True


    # 쿼리 정의하기
    while(flag):
        flag = False
        queries = ""

        for text in texts:
            #queries += "\"{}\"의 키워드를 한글로 알려줘. ".format(text)
            queries += "\"{}\"를 한단어로 요약해줘. ".format(text)

        # 모델 - GPT 3.5 Turbo 선택
        model = "gpt-3.5-turbo"

        # 메시지 설정하기
        messages = [
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": queries}
        ]

        # ChatGPT API 호출하기
        response = openai.ChatCompletion.create(
            model=model,
            messages=messages
        )
        answer = response['choices'][0]['message']['content']

        # 키워드 정리
        keywords = list(set(okt.nouns(answer)))

        for keyword in keywords:
            if len(keyword) <= 1:
                flag = True
                break

        if len(keywords) == 0:
            flag = True

        if flag:
            print("sleep...")
            time.sleep(1)

    for keyword in keywords:
        if keyword not in wrong_word:
            result["keyword"][keyword] = 10

    vmax = 0

    for key, value in result["keyword"].items():
        vmax = max(value, vmax)

    for key, value in result["keyword"].items():
        result["keyword"][key] = round(value / vmax, 5)

    end = time.time()
    result["time"] = round(end - start, 5)

    return result

def CUSTOM_MODEL_01(KoWR_result, Bert_result, SenNouns_result, GPT_result, wrong_word, V):
    result = {"time":{}, "keyword":{}}
    start = time.time()

    # Bert Base
    for key, value in Bert_result["keyword"].items():
        if key in list(KoWR_result["keyword"].keys()):
            result["keyword"][key] = float(value)

    # Add KoWR
    for key, value in KoWR_result["keyword"].items():
        if key in result["keyword"]:
            result["keyword"][key] += (value * V)

    # Delete frequent noun
    noun = 0

    for key, value in SenNouns_result["keyword"].items():
        if value == 1:
            noun = key
            break

    if noun != 0:
        for key, value in result["keyword"].items():
            if key == noun:
                del result["keyword"][key]
                break

    # desc order
    D1 = sorted(result["keyword"].items(), key=lambda x: x[1], reverse=True)
    result["keyword"] = {}

    for d1 in D1:
        # delete wrong word
        if d1[0] not in wrong_word:
            result["keyword"][d1[0]] = d1[1]

    end = time.time()
    result["time"] = round(end - start, 5)

    return result

def WriteExcel(path, input_name, output_name):
    input_path = "{}/{}".format(path, input_name)
    output_path = "{}/{}".format(path, output_name)

    workbook_input = load_workbook(filename=input_path)
    worksheet_input = workbook_input['Sheet1']

    workbook_output = Workbook()
    worksheet_output = workbook_output.create_sheet("Sheet1")
    worksheet_output = workbook_output.active

    worksheet_output['A1'] = "인덱스"
    worksheet_output['B1'] = "회사 이름"
    worksheet_output['C1'] = "메인 키워드"
    worksheet_output['D1'] = "서브 키워드"

    # worksheet["A1"].value

    result = []
    output_idx = 2

    for i in range(1, 10000):
        code_val = worksheet_input["A{}".format(i)].value
        code_name = worksheet_input["B{}".format(i)].value
        code_info = worksheet_input["C{}".format(i)].value

        if code_val is not None and int(code_val) == 0:
            break

        if code_info is not None:
            if code_val is not None:
                result.append([int(code_val), code_name, [code_info.replace("\n", "")]])

            else:
                result[-1][2].append(code_info)

    for res in result:
        print(res[1])

        KoWR_result = KoWordRankModule(res[2])
        Bert_result = KeyBert(res[2])
        SenNouns_result = SenNouns(res[2])
        GPT_result = GPT(res[2])
        custom_model_01 = CUSTOM_MODEL_01(KoWR_result, Bert_result, SenNouns_result, GPT_result,
                                          wrong_word, 0.5)

        main_keys = []
        sub_keys = []

        print("Main keyword")
        for key, value in GPT_result["keyword"].items():
            main_keys.append(key)
            main_all.add(key)
        print(main_keys)

        print("Sub keyword")
        for key, value in custom_model_01["keyword"].items():
            sub_keys.append(key)
            sub_all.add(key)
        print(sub_keys)

        # write
        worksheet_output["A{}".format(output_idx)] = str(res[0])
        worksheet_output["B{}".format(output_idx)] = str(res[1])

        midx = output_idx
        for i in range(len(main_keys)):
            worksheet_output["C{}".format(midx)] = str(main_keys[i])
            midx += 1

        sidx = output_idx
        for i in range(len(sub_keys)):
            worksheet_output["D{}".format(sidx)] = str(sub_keys[i])
            sidx += 1

        output_idx += max(len(main_keys), len(sub_keys))

        print("END")

    workbook_output.save(output_path)
    print(len(main_all))
    print(main_all)

    print(len(sub_all))
    print(sub_all)

if __name__ == '__main__':
    path = "../data/info01.txt"
    beginTime = time.time()

    WriteExcel("C:/Users/SSAFY/Desktop/excel", "input11_20.xlsx", "output11_20.xlsx")

    endTime = time.time()
    print(round(endTime - beginTime))